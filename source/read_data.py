import pandas as pd
import os
from openpyxl import load_workbook
from source.utils import list_files, solve_folder
import pprint

class MockDB:
    '''Mocks a database by reading the excel files extracted 
    by the IT department'''


    DATA_DIR = solve_folder('db_data', 'original_data')
    FILE_EXTENSIONS = '.xlsx'
    
    def __init__(self):
        
        self.files = self.get_xl_files()
        self.schema = self.get_schema()
        
    def get_xl_files(self):
        
        xls = list_files(self.DATA_DIR, extension=self.FILE_EXTENSIONS)
        
        return xls
    
    def get_col_names(self, fname):
        
        wb = load_workbook(fname, read_only=True)
        active_ws = wb.active
        
        rows = active_ws.iter_rows(min_row=1, max_row=1)
        first_row = next(rows)
        headings = [c.value for c in first_row]
        wb.close()
        
        return headings
    
    def get_table_name(self, xl_file_path):
        
        fname = os.path.split(xl_file_path)[-1]
        table_name = fname.replace(self.FILE_EXTENSIONS, '')
        
        #must remove date of extraction
        table_name = ''.join([char for char in 
                             table_name if not char.isdigit()])
        
        return table_name
    
    def get_schema(self):
        
        xlsx = self.files
        
        schema = {}
        for xl in xlsx:
            table = self.get_table_name(xl)
            columns = self.get_col_names(xl)
            schema[table] = columns
        
        return schema
    
    def fname_from_tablename(self, table):
        
        for file in self.files:
            if table in file:
                return file

    def __repr__(self):

        return 'MockDB()'
    
    def __str__(self):
        
        return pprint.pformat(self.schema)
    
    def __getitem__(self, table_name):
        
        file = self.fname_from_tablename(table_name)
        
        return pd.read_excel(file)